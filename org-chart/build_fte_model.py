#!/usr/bin/env python3
"""AWDS Construction-Stage FTE / resourcing & cost model — a rate for every title."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY = "0B2E63"; BLUE = "1F5FD6"; LIGHT = "EAF1FC"; SUB = "DCE7F8"
MINT = "E4F6F0"; ROSE = "FBE9F2"; GREY = "F2F5FA"; GOLD = "FFF3D6"

f_title = Font(name="Calibri", size=16, bold=True, color=NAVY)
f_sub   = Font(name="Calibri", size=10, italic=True, color="5B6B8C")
f_hdr   = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
f_bold  = Font(name="Calibri", size=10, bold=True, color=NAVY)
f_norm  = Font(name="Calibri", size=10, color="0F1B3D")
f_note  = Font(name="Calibri", size=9, italic=True, color="5B6B8C")
f_input = Font(name="Calibri", size=11, bold=True, color="1F5FD6")
f_white = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
f_link  = Font(name="Calibri", size=10, color="1F5FD6")

fill_hdr = PatternFill("solid", fgColor=NAVY)
fill_sub = PatternFill("solid", fgColor=SUB)
fill_light = PatternFill("solid", fgColor=LIGHT)
fill_grey = PatternFill("solid", fgColor=GREY)
fill_tot = PatternFill("solid", fgColor=BLUE)
fill_input = PatternFill("solid", fgColor=GOLD)

thin = Side(style="thin", color="C5D2EA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
EUR = '"€"#,##0'
HRS = "'Rates'!$C$5"

# ---------------------------------------------------------------------
# Master data: (name, headcount, utilisation, duration, rate €/hr, basis)
# ---------------------------------------------------------------------
mgmt = [
    ("AWDS Project Director", 1, 0.25, 36, 160, "Part-time senior oversight & governance"),
    ("AWDS Project Management Lead", 1, 1.00, 36, 130, "Full-time; leads delivery across all packages"),
    ("AWDS Deputy Project Manager", 1, 1.00, 36, 110, "Full-time; daily operations & coordination"),
    ("AWDS Design Management Lead", 1, 1.00, 36, 120, "Full-time; design integration & reachback mgmt"),
    ("AWDS PMO Lead", 1, 1.00, 36, 105, "Full-time; controls, reporting, information mgmt"),
    ("AWDS Technical Delivery Lead", 1, 1.00, 36, 120, "Full-time; technical assurance & verification"),
    ("AWDS Commercial Lead", 1, 1.00, 36, 120, "Full-time; NEC commercial & change"),
    ("AWDS Cost Management Lead", 1, 1.00, 36, 110, "Full-time; cost control & forecasting"),
]
cm = [
    ("AWDS Contract Manager (one per package)", 16, 1.00, 18, 100,
     "Package PM: RFIs, contractor & utility-owner meetings, design-proposal review, as-builts, NEC admin"),
]
contract_support = [
    ("Contracts / Technical Engineer", 8, 0.75, 18, 60, "Reviews contractor design proposals, raises/closes RFIs (~1 per 2 packages)"),
    ("Document Controller / As-Built Coordinator", 4, 0.60, 18, 40, "Manages & updates as-built drawings and CDE (~1 per 4 packages)"),
    ("Quantity Surveyor / Commercial Support", 4, 0.60, 18, 65, "NEC change, valuations and payment support (~1 per 4 packages)"),
    ("Utilities Liaison Coordinator", 2, 0.50, 18, 55, "Meetings & coordination with utility asset owners"),
    ("CEMP / Environmental Coordinator", 2, 0.40, 18, 55, "Environmental compliance and CEMP across packages"),
    ("Planner / Scheduler (package)", 2, 0.50, 18, 60, "Programme tracking at package level"),
]
project_support = [
    ("BIM Manager", 1, 0.50, 18, 70, "Shared across programme"),
    ("Deputy BIM Manager", 1, 0.50, 18, 50, ""),
    ("Information Manager", 1, 0.50, 18, 65, ""),
    ("Deputy Information Manager", 1, 0.40, 18, 48, ""),
    ("Senior GIS Consultant", 1, 0.30, 18, 65, "On-demand"),
    ("Senior GIS Analyst", 1, 0.40, 18, 48, ""),
    ("Surveys Coordinator", 1, 0.30, 18, 48, "Ad-hoc survey requests"),
    ("CDE Document Controller (project)", 1, 0.70, 18, 40, "Project-wide CDE"),
    ("BusConnects Interface", 1, 0.20, 18, 60, "Interface management"),
    ("DAA Interface", 1, 0.20, 18, 60, "Interface management"),
    ("UAO Engagement Team", 2, 0.30, 18, 52, "Utility/asset-owner engagement"),
    ("Administrator", 2, 0.80, 18, 32, "Team & document admin"),
    ("Health & Safety Advisor", 1, 0.50, 18, 60, "CDM / safety in design support"),
    ("QA / QC Manager", 1, 0.50, 18, 65, "Quality assurance"),
    ("Risk Manager", 1, 0.30, 18, 70, ""),
    ("Project Controls Analyst", 1, 0.60, 18, 55, "Reporting & dashboards"),
    ("HR & People Coordinator", 1, 0.25, 18, 42, ""),
    ("IT / Systems Support", 1, 0.25, 18, 42, ""),
    ("Communications & Stakeholder Liaison", 1, 0.30, 18, 52, ""),
    ("Training & Competency Coordinator", 1, 0.20, 18, 45, ""),
    ("Office Manager", 1, 0.60, 18, 42, ""),
]
designers = [
    ("Design Management / Coordination", 1, 0.30, 18, 110, "Coordinates reachback requests"),
    ("Utilities South", 3, 0.25, 18, 95, "Reachback — design queries & proposal review"),
    ("Sustainability", 2, 0.15, 18, 90, "On-demand"),
    ("Land Access", 2, 0.20, 18, 90, "On-demand"),
    ("Structures", 2, 0.25, 18, 100, "Design review & RFI support"),
    ("Heritage", 1, 0.10, 18, 90, "On-demand"),
    ("Cellars & Playing Pitches", 1, 0.10, 18, 90, "On-demand"),
    ("Archaeology", 1, 0.15, 18, 90, "On-demand"),
    ("Biodiversity", 1, 0.10, 18, 90, "On-demand"),
    ("Env. Monitoring", 2, 0.20, 18, 85, "Periodic monitoring"),
]

wb = openpyxl.Workbook()

def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = f_hdr; cell.fill = fill_hdr; cell.alignment = center; cell.border = border

# =====================================================================
# SHEET: Assumptions
# =====================================================================
wa = wb.active; wa.title = "Assumptions"; wa.sheet_view.showGridLines = False
wa["A1"] = "AWDS MetroLink — Construction Stage Resourcing, FTE & Cost Model"; wa["A1"].font = f_title
wa["A2"] = "Edit the gold cells on the 'Rates' tab (a rate for every title) — all sheets and the total cost update automatically."; wa["A2"].font = f_sub
arows = [
    ("", ""), ("KEY ASSUMPTIONS", ""),
    ("Construction-stage durations", ""),
    ("   • Management team", "36 months"),
    ("   • Contract Managers (and their delivery team)", "18 months"),
    ("   • Support staff", "18 months"),
    ("   • Designers (reachback)", "18 months"),
    ("Number of contract packages / Contract Managers", "16  (one per package — see note 1)"),
    ("Rates & hours", "See the 'Rates' tab — one editable rate per title."),
    ("", ""), ("DEFINITIONS", ""),
    ("FTE", "1.0 FTE = one full-time-equivalent person at 100% utilisation."),
    ("FTE (modelled)", "= Headcount × Utilisation %"),
    ("Effort (person-months)", "= FTE × Duration (months)"),
    ("Cost (€)", "= Person-months × Hours-per-month × the title's rate (from 'Rates' tab)"),
    ("", ""), ("FTE vs NON-FTE TREATMENT", ""),
    ("Management team & Contract Managers", "Dedicated FTE."),
    ("Support staff & Designers", "Shared / part-time, counted by utilisation."),
    ("", ""), ("SCOPE CONTEXT (drives the team make-up)", ""),
    ("At construction stage the Contract Managers and their teams are NOT site supervision. Their remit is:", ""),
    ("   • Raising / tracking / closing RFIs", ""),
    ("   • Reviewing contractor design proposals", ""),
    ("   • Progress, design and coordination meetings; occasional site visits", ""),
    ("   • Meetings with utility asset owners", ""),
    ("   • Updating and managing the as-built drawings / CDE", ""),
    ("   • NEC contract administration, change and reporting", ""),
    ("", ""), ("NOTES", ""),
    ("1.  Contract Manager headcount (16) is from the org chart; some packages combine under one CM —", ""),
    ("     edit the headcount on sheet 1 to suit.", ""),
    ("2.  Cost is labour only — excludes expenses, fee/overhead uplift and inflation.", ""),
    ("3.  Per-title rates are planning assumptions — refine on the 'Rates' tab.", ""),
]
r = 4
for label, val in arows:
    wa.cell(row=r, column=1, value=label); wa.cell(row=r, column=3, value=val)
    if label in ("KEY ASSUMPTIONS","DEFINITIONS","FTE vs NON-FTE TREATMENT","SCOPE CONTEXT (drives the team make-up)","NOTES"):
        wa.cell(row=r, column=1).font = f_bold
        for c in range(1,7): wa.cell(row=r, column=c).fill = fill_sub
    else:
        wa.cell(row=r, column=1).font = f_norm; wa.cell(row=r, column=3).font = f_norm
    wa.cell(row=r, column=3).alignment = left
    r += 1
wa.column_dimensions["A"].width = 52; wa.column_dimensions["B"].width = 2; wa.column_dimensions["C"].width = 66

# =====================================================================
# SHEET: Rates  (one rate per title — source of truth)
# =====================================================================
wr = wb.create_sheet("Rates"); wr.sheet_view.showGridLines = False
wr["A1"] = "Rates & Cost Inputs — a rate for every title"; wr["A1"].font = f_title
wr["A2"] = "Edit any gold cell. Every sheet and the total cost recalculate automatically."; wr["A2"].font = f_sub
wr["A4"] = "GLOBAL INPUT"; wr["A4"].font = f_bold
for c in range(1,5): wr.cell(row=4, column=c).fill = fill_sub
wr["A5"] = "Hours per person-month"; wr["A5"].font = f_bold
ci = wr.cell(row=5, column=3, value=160); ci.font = f_input; ci.fill = fill_input; ci.alignment = center; ci.number_format = "0"; ci.border = border
wr["D5"] = "Converts person-months to chargeable hours"; wr["D5"].font = f_note
wr["A7"] = "Indicative total construction-stage cost (€)"; wr["A7"].font = f_bold
tc = wr.cell(row=7, column=3, value="='Summary'!$H$10"); tc.font = f_white; tc.fill = fill_tot; tc.alignment = center; tc.number_format = EUR; tc.border = border
wr["D7"] = "Live — labour only (see Assumptions)"; wr["D7"].font = f_note

# rate table header
hr = 9
wr.cell(row=hr, column=1, value="Role / Title"); wr.cell(row=hr, column=3, value="Rate (€/hr)"); wr.cell(row=hr, column=4, value="Category")
for c in (1,3,4):
    cell = wr.cell(row=hr, column=c); cell.font = f_hdr; cell.fill = fill_hdr; cell.alignment = center; cell.border = border

rate_cell = {}   # title -> 'Rates'!$C$row
groups = [
    ("MANAGEMENT TEAM", mgmt), ("CONTRACT MANAGERS", cm),
    ("CONTRACT DELIVERY SUPPORT", contract_support),
    ("PROJECT SUPPORT — Interface & Back-Office", project_support),
    ("DESIGNERS (reachback)", designers),
]
row = hr + 1
for gname, items in groups:
    wr.cell(row=row, column=1, value=gname).font = f_bold
    for c in range(1,5): wr.cell(row=row, column=c).fill = fill_sub; wr.cell(row=row, column=c).border = border
    row += 1
    for name, hc, util, dur, rate, basis in items:
        wr.cell(row=row, column=1, value=name).font = f_norm
        rc = wr.cell(row=row, column=3, value=rate); rc.font = f_input; rc.fill = fill_input; rc.alignment = center; rc.number_format = EUR
        wr.cell(row=row, column=4, value=gname).font = f_note
        for c in (1,3,4): wr.cell(row=row, column=c).border = border
        rate_cell[name] = f"'Rates'!$C${row}"
        row += 1
wr.column_dimensions["A"].width = 44; wr.column_dimensions["B"].width = 2
wr.column_dimensions["C"].width = 13; wr.column_dimensions["D"].width = 40
wr.freeze_panes = "A10"

# =====================================================================
# Builder: resourcing + per-title rate + cost
# =====================================================================
HEAD = ["Role", "Headcount", "Utilisation %", "FTE", "Duration (months)",
        "Effort (person-months)", "Effort (person-years)", "Rate (€/hr)", "Cost (€)", "Basis / notes"]

def build_sheet(title, intro, sections):
    ws = wb.create_sheet(title); ws.sheet_view.showGridLines = False
    ws["A1"] = intro[0]; ws["A1"].font = f_title
    ws["A2"] = intro[1]; ws["A2"].font = f_sub
    hr = 4
    for c, h in enumerate(HEAD, start=1): ws.cell(row=hr, column=c, value=h)
    style_header_row(ws, hr, len(HEAD))
    row = hr + 1; sub_rows = []
    for sec_name, sec_fill, items in sections:
        ws.cell(row=row, column=1, value=sec_name).font = f_bold
        for c in range(1, len(HEAD)+1): ws.cell(row=row, column=c).fill = fill_sub; ws.cell(row=row, column=c).border = border
        row += 1; first = row
        for name, hc, util, dur, rate, basis in items:
            rcell = rate_cell[name]
            ws.cell(row=row, column=1, value=name).font = f_norm; ws.cell(row=row, column=1).alignment = left
            ws.cell(row=row, column=2, value=hc).alignment = center
            ws.cell(row=row, column=3, value=util).number_format = "0%"; ws.cell(row=row, column=3).alignment = center
            ws.cell(row=row, column=4, value=f"=B{row}*C{row}").number_format = "0.00"; ws.cell(row=row, column=4).alignment = center
            ws.cell(row=row, column=5, value=dur).alignment = center
            ws.cell(row=row, column=6, value=f"=D{row}*E{row}").number_format = "0.0"; ws.cell(row=row, column=6).alignment = center
            ws.cell(row=row, column=7, value=f"=F{row}/12").number_format = "0.0"; ws.cell(row=row, column=7).alignment = center
            rcl = ws.cell(row=row, column=8, value=f"={rcell}"); rcl.number_format = EUR; rcl.alignment = center; rcl.font = f_link
            ws.cell(row=row, column=9, value=f"=F{row}*{HRS}*{rcell}").number_format = EUR; ws.cell(row=row, column=9).alignment = center
            ws.cell(row=row, column=10, value=basis).font = f_note; ws.cell(row=row, column=10).alignment = left
            for c in range(1, len(HEAD)+1):
                ws.cell(row=row, column=c).border = border
                if sec_fill: ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=sec_fill)
            row += 1
        last = row - 1
        ws.cell(row=row, column=1, value=f"   Subtotal — {sec_name}").font = f_bold
        ws.cell(row=row, column=2, value=f"=SUM(B{first}:B{last})").alignment = center
        ws.cell(row=row, column=4, value=f"=SUM(D{first}:D{last})").number_format = "0.00"; ws.cell(row=row, column=4).alignment = center
        ws.cell(row=row, column=6, value=f"=SUM(F{first}:F{last})").number_format = "0.0"; ws.cell(row=row, column=6).alignment = center
        ws.cell(row=row, column=7, value=f"=SUM(G{first}:G{last})").number_format = "0.0"; ws.cell(row=row, column=7).alignment = center
        ws.cell(row=row, column=9, value=f"=SUM(I{first}:I{last})").number_format = EUR; ws.cell(row=row, column=9).alignment = center
        for c in range(1, len(HEAD)+1):
            ws.cell(row=row, column=c).fill = fill_light; ws.cell(row=row, column=c).border = border
            if c != 1: ws.cell(row=row, column=c).font = f_bold
        sub_rows.append(row); row += 2
    ws.cell(row=row, column=1, value="TOTAL").font = f_white
    ws.cell(row=row, column=2, value="=" + "+".join(f"B{x}" for x in sub_rows)).alignment = center
    ws.cell(row=row, column=4, value="=" + "+".join(f"D{x}" for x in sub_rows)).number_format = "0.00"
    ws.cell(row=row, column=6, value="=" + "+".join(f"F{x}" for x in sub_rows)).number_format = "0.0"
    ws.cell(row=row, column=7, value="=" + "+".join(f"G{x}" for x in sub_rows)).number_format = "0.0"
    ws.cell(row=row, column=9, value="=" + "+".join(f"I{x}" for x in sub_rows)).number_format = EUR
    for c in range(1, len(HEAD)+1):
        ws.cell(row=row, column=c).fill = fill_tot; ws.cell(row=row, column=c).border = border
        if c != 1: ws.cell(row=row, column=c).font = f_white; ws.cell(row=row, column=c).alignment = center
    widths = [42, 11, 13, 8, 16, 19, 18, 12, 14, 46]
    for i, w in enumerate(widths, start=1): ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"

build_sheet("1 · Mgmt & Contract Mgrs",
    ("1 · Management Team & Contract Managers — FTE",
     "Dedicated full-time-equivalent resource. Each title's rate is set on the 'Rates' tab."),
    [("MANAGEMENT TEAM  (36 months)", None, mgmt), ("CONTRACT MANAGERS  (18 months)", None, cm)])
build_sheet("2 · Support Staff",
    ("2 · Support Staff — utilisation-based (non-FTE)",
     "Shared / part-time resources. Each title's rate is set on the 'Rates' tab."),
    [("CONTRACT DELIVERY SUPPORT  (team around the Contract Managers, 18 months)", ROSE, contract_support),
     ("PROJECT SUPPORT — Interface & Back-Office  (18 months)", MINT, project_support)])
build_sheet("3 · Designers",
    ("3 · Designers — reachback, utilisation-based (non-FTE)",
     "Design discipline reachback team. Each title's rate is set on the 'Rates' tab."),
    [("DESIGN REACHBACK TEAM  (18 months)", MINT, designers)])

# =====================================================================
# SHEET: Summary
# =====================================================================
ws = wb.create_sheet("Summary"); ws.sheet_view.showGridLines = False
ws["A1"] = "Construction Stage — Resource & Cost Summary"; ws["A1"].font = f_title
ws["A2"] = "Roll-up of all categories. Cost is driven by the per-title rates on the 'Rates' tab."; ws["A2"].font = f_sub
heads = ["Category", "Type", "Headcount", "FTE (modelled)", "Duration (months)",
         "Effort (person-months)", "Effort (person-years)", "Cost (€)"]
hr = 4
for c, h in enumerate(heads, start=1): ws.cell(row=hr, column=c, value=h)
style_header_row(ws, hr, len(heads))
S1 = "'1 · Mgmt & Contract Mgrs'"; S2 = "'2 · Support Staff'"; S3 = "'3 · Designers'"
data = [
    ("Management Team", "FTE (dedicated)", f"{S1}!B14", f"{S1}!D14", 36, f"{S1}!F14", f"{S1}!G14", f"{S1}!I14"),
    ("Contract Managers", "FTE (dedicated)", f"{S1}!B18", f"{S1}!D18", 18, f"{S1}!F18", f"{S1}!G18", f"{S1}!I18"),
    ("Contract Delivery Support", "Utilisation", f"{S2}!B12", f"{S2}!D12", 18, f"{S2}!F12", f"{S2}!G12", f"{S2}!I12"),
    ("Project Support (Interface & Back-Office)", "Utilisation", f"{S2}!B36", f"{S2}!D36", 18, f"{S2}!F36", f"{S2}!G36", f"{S2}!I36"),
    ("Designers (reachback)", "Utilisation", f"{S3}!B16", f"{S3}!D16", 18, f"{S3}!F16", f"{S3}!G16", f"{S3}!I16"),
]
row = hr + 1; first = row
for cat, typ, hc, fte, dur, pm, py, cost in data:
    ws.cell(row=row, column=1, value=cat).font = f_norm
    ws.cell(row=row, column=2, value=typ).font = f_norm
    ws.cell(row=row, column=3, value="=" + hc).alignment = center
    ws.cell(row=row, column=4, value="=" + fte).number_format = "0.00"; ws.cell(row=row, column=4).alignment = center
    ws.cell(row=row, column=5, value=dur).alignment = center
    ws.cell(row=row, column=6, value="=" + pm).number_format = "0.0"; ws.cell(row=row, column=6).alignment = center
    ws.cell(row=row, column=7, value="=" + py).number_format = "0.0"; ws.cell(row=row, column=7).alignment = center
    ws.cell(row=row, column=8, value="=" + cost).number_format = EUR; ws.cell(row=row, column=8).alignment = center
    for c in range(1, len(heads)+1):
        ws.cell(row=row, column=c).border = border
        if row % 2 == 0: ws.cell(row=row, column=c).fill = fill_grey
    row += 1
last = row - 1
ws.cell(row=row, column=1, value="TOTAL — Construction Stage").font = f_white
ws.cell(row=row, column=3, value=f"=SUM(C{first}:C{last})").alignment = center
ws.cell(row=row, column=4, value=f"=SUM(D{first}:D{last})").number_format = "0.00"; ws.cell(row=row, column=4).alignment = center
ws.cell(row=row, column=6, value=f"=SUM(F{first}:F{last})").number_format = "0.0"; ws.cell(row=row, column=6).alignment = center
ws.cell(row=row, column=7, value=f"=SUM(G{first}:G{last})").number_format = "0.0"; ws.cell(row=row, column=7).alignment = center
ws.cell(row=row, column=8, value=f"=SUM(H{first}:H{last})").number_format = EUR; ws.cell(row=row, column=8).alignment = center
for c in range(1, len(heads)+1):
    ws.cell(row=row, column=c).fill = fill_tot; ws.cell(row=row, column=c).border = border
    if c != 1: ws.cell(row=row, column=c).font = f_white
row += 2
ws.cell(row=row, column=1, value="FTE (Management + Contract Managers)").font = f_bold
ws.cell(row=row, column=4, value=f"=D{first}+D{first+1}").number_format = "0.00"; ws.cell(row=row,column=4).alignment=center; ws.cell(row=row,column=4).fill = fill_sub
row += 1
ws.cell(row=row, column=1, value="FTE-equivalent (Support + Designers)").font = f_bold
ws.cell(row=row, column=4, value=f"=D{first+2}+D{first+3}+D{first+4}").number_format = "0.00"; ws.cell(row=row,column=4).alignment=center; ws.cell(row=row,column=4).fill = fill_sub
for i, w in enumerate([40, 18, 12, 16, 18, 22, 20, 16], start=1): ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A5"

order = ["Assumptions", "Rates", "Summary", "1 · Mgmt & Contract Mgrs", "2 · Support Staff", "3 · Designers"]
wb._sheets.sort(key=lambda s: order.index(s.title))

out = "/home/user/phd-tracker/org-chart/AWDS_Construction_Stage_FTE_Model.xlsx"
wb.save(out)
print("saved", out, "| sheets:", wb.sheetnames, "| titles priced:", len(rate_cell))
